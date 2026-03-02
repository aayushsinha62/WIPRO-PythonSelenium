from openpyxl import load_workbook

def read_test_data(file_path, sheet_name):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    data_list = []
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if all(cell is None for cell in row):
            continue

        data = dict(zip(headers, row))
        data_list.append(data)

    return data_list